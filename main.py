
import os
import re
import sys
import argparse
import datetime
import ipaddress
from collections import Counter, defaultdict
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

VT_DOMAIN_URL = "https://www.virustotal.com/api/v3/domains/{}"
VT_RELATIONS_URL = "https://www.virustotal.com/api/v3/domains/{}/relationships/{}"
VT_HISTORICAL_WHOIS_URL = "https://www.virustotal.com/api/v3/domains/{}/historical_whois"

CLUSTER_COLORS = [
    "E2EFDA",
    "DDEBF7",
    "FCE4D6",
    "FFF2CC",
    "E8D7F1",
    "D9E1F2",
    "E2F0D9",
    "F8CBAD",
    "D0CECE",
]

SCORE_THRESHOLD = 30


def clean_domain(raw_domain: str) -> str:
    """Очистка домена от дефангинга, протоколов и URI."""
    d = raw_domain.strip().lower()
    d = re.sub(r"\[\.\]|\(\.\)", ".", d)
    d = re.sub(r"^https?://", "", d)
    d = d.split("/")[0]
    return d


def defang_domain(domain: str) -> str:
    """Дефангинг домена для безопасного экспорта в отчет."""
    return domain.replace(".", "[.]")


def normalize_registrar_name(raw: str) -> str:
    """Очистка и нормализация имени регистратора."""
    if not raw or raw == "Н/Д":
        return "Н/Д"

    val = raw.strip()
    if "http" in val or "." in val:
        val = re.sub(r"^https?://", "", val, flags=re.IGNORECASE)
        val = val.split("/")[0].split(":")[0].strip()
        val = re.sub(r"^(?:whois|www|rdap|cp|ns\d*|web)\.", "", val, flags=re.IGNORECASE)

    val = val.strip(" \"',;")
    return val if val else "Н/Д"


def extract_registrar_from_text(whois_text: str) -> str:
    """Точный парсинг регистратора из сырого текста WHOIS."""
    if not whois_text:
        return "Н/Д"

    url_match = re.search(
        r"(?:Registrar URL|Registrar\s+Web|Registrar\s+Homepage):?\s*(https?://[^\s\r\n]+|[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})",
        whois_text, re.IGNORECASE)
    if url_match:
        return normalize_registrar_name(url_match.group(1))

    for line in whois_text.splitlines():
        line_clean = line.strip()
        if re.match(r"^(?:Registrar|Sponsoring Registrar|Registrar Name):\s*(.+)$", line_clean, re.IGNORECASE):
            if any(bad in line_clean.lower() for bad in
                   ["iana", " id", "abuse", "whois server", "url", "phone", "email"]):
                continue
            val = re.sub(r"^(?:Registrar|Sponsoring Registrar|Registrar Name):\s*", "", line_clean,
                         flags=re.IGNORECASE).strip()
            if val and val.lower() != "n/a":
                return normalize_registrar_name(val)

    return "Н/Д"


def extract_country_from_dict_or_text(whois_map: dict, whois_text: str) -> str:
    """Извлечение страны"""
    if whois_map and isinstance(whois_map, dict):
        for k, v in whois_map.items():
            k_lower = str(k).lower().replace(" ", "_").replace("-", "_")
            if "registrant" in k_lower and "country" in k_lower and v:
                val = str(v).strip()
                if val and val.lower() not in ["null", "n/a", "none", "н/д"]:
                    return val

        for k, v in whois_map.items():
            k_lower = str(k).lower()
            if k_lower in ["country", "registrant_country", "registrant-country"] and v:
                val = str(v).strip()
                if val and val.lower() not in ["null", "n/a", "none", "н/д"]:
                    return val

    if whois_text:
        match = re.search(r"Registrant\s+(?:Contact\s+)?Country:?\s*([^\r\n]+)", whois_text, re.IGNORECASE)
        if match:
            val = match.group(1).strip(" \"',;")
            if val and val.lower() not in ["null", "n/a", "none", "н/д"]:
                return val

        match_alt = re.search(r"(?:registrant_country|Country):?\s*([A-Za-z\s]+)", whois_text, re.IGNORECASE)
        if match_alt:
            val = match_alt.group(1).strip(" \"',;")
            if val and val.lower() not in ["null", "n/a", "none", "н/д", "state"]:
                return val

    return "Н/Д"


def fetch_historical_whois_data(domain: str, headers: dict) -> tuple[str, str]:
    """Поиск страны и регистратора в historical_whois."""
    hist_country = "Н/Д"
    hist_registrar = "Н/Д"

    try:
        resp = requests.get(VT_HISTORICAL_WHOIS_URL.format(domain), headers=headers, timeout=10)
        if resp.status_code == 200:
            records = resp.json().get("data", [])
            records.sort(key=lambda x: x.get("attributes", {}).get("timestamp", 0), reverse=True)

            for rec in records:
                attrs = rec.get("attributes", {})
                w_map = attrs.get("whois_map", {})
                raw_text = w_map.get("raw", "") or attrs.get("raw", "") or ""

                if hist_country == "Н/Д":
                    found_c = extract_country_from_dict_or_text(w_map, raw_text)
                    if found_c != "Н/Д":
                        hist_country = found_c

                if hist_registrar == "Н/Д":
                    found_r = extract_registrar_from_text(raw_text)
                    if found_r != "Н/Д":
                        hist_registrar = found_r

                if hist_country != "Н/Д" and hist_registrar != "Н/Д":
                    break
    except Exception as e:
        print(f"[-] Ошибка historical_whois для {domain}: {e}")

    return hist_registrar, hist_country


def format_rname_to_email(rname: str) -> str:
    """Преобразование rname из SOA в email."""
    if not rname or rname == "Н/Д":
        return "Н/Д"
    rname = rname.rstrip(".")
    if "@" in rname:
        return rname
    rname = re.sub(r'(?<!\\)\.', '@', rname, count=1)
    rname = rname.replace('\\.', '.')
    return rname


def compute_subnet(ip_str: str) -> str:
    """Вычисление /24 подсети для IPv4."""
    try:
        ip = ipaddress.ip_address(ip_str.strip())
        if isinstance(ip, ipaddress.IPv4Address):
            net = ipaddress.ip_network(f"{ip}/24", strict=False)
            return str(net)
    except ValueError:
        pass
    return "Н/Д"


def fetch_virustotal_data(domain: str, api_key: str) -> dict:
    """Сбор пассивных индикаторов"""
    headers = {"x-apikey": api_key}
    result = {
        "registrar": "Н/Д", "reg_country": "Н/Д", "creation_date": "Н/Д",
        "ns_records": "Н/Д", "mx_records": "Н/Д", "soa_rname": "Н/Д",
        "passive_ips": "Н/Д", "subnets": "Н/Д", "hist_ips_str": "Н/Д",
        "ssl_sha256": "Н/Д", "hist_ssl_str": "Н/Д", "ssl_san": "Н/Д",
        "ssl_valid_from": "Н/Д", "vt_relations": "Н/Д",
        "creation_dt": None,
        "raw_ns": [], "raw_mx": [], "raw_ips": [], "raw_subnets": [],
        "raw_hist_ips": [], "raw_hist_ssl": [], "raw_hashes": []
    }

    try:
        resp = requests.get(VT_DOMAIN_URL.format(domain), headers=headers, timeout=15)
        if resp.status_code == 200:
            data = resp.json().get("data", {}).get("attributes", {})
            whois_text = data.get("whois", "")
            whois_map = data.get("whois_map", {})

            registrar = extract_registrar_from_text(whois_text)
            country = extract_country_from_dict_or_text(whois_map, whois_text)

            if registrar == "Н/Д" or country == "Н/Д":
                h_reg, h_country = fetch_historical_whois_data(domain, headers)
                if registrar == "Н/Д": registrar = h_reg
                if country == "Н/Д": country = h_country

            result["registrar"] = registrar
            result["reg_country"] = country

            c_date_ts = data.get("creation_date")
            if c_date_ts:
                dt = datetime.datetime.utcfromtimestamp(c_date_ts)
                result["creation_date"] = dt.strftime("%Y-%m-%d")
                result["creation_dt"] = dt

            dns_records = data.get("last_dns_records", [])
            for rec in dns_records:
                rtype = rec.get("type")
                val = rec.get("value", "").strip()
                if rtype == "A" and val:
                    result["raw_ips"].append(val)
                elif rtype == "NS" and val:
                    result["raw_ns"].append(val.rstrip("."))
                elif rtype == "MX" and val:
                    result["raw_mx"].append(val)
                elif rtype == "SOA" and result["soa_rname"] == "Н/Д":
                    soa_rname_val = rec.get("rname", "")
                    if soa_rname_val:
                        result["soa_rname"] = format_rname_to_email(soa_rname_val)

            if result["raw_ips"]:
                unique_ips = list(dict.fromkeys(result["raw_ips"]))
                result["raw_ips"] = unique_ips
                result["passive_ips"] = ", ".join(unique_ips)

                subnets = list(dict.fromkeys([compute_subnet(ip) for ip in unique_ips if compute_subnet(ip) != "Н/Д"]))
                result["raw_subnets"] = subnets
                result["subnets"] = ", ".join(subnets) if subnets else "Н/Д"

            if result["raw_ns"]:
                result["raw_ns"] = list(dict.fromkeys(result["raw_ns"]))
                result["ns_records"] = ", ".join(result["raw_ns"])

            if result["raw_mx"]:
                result["raw_mx"] = list(dict.fromkeys(result["raw_mx"]))
                result["mx_records"] = ", ".join(result["raw_mx"])

            cert = data.get("last_https_certificate")
            if cert:
                result["ssl_sha256"] = cert.get("thumbprint_sha256") or cert.get("sha256") or "Н/Д"
                extensions = cert.get("extensions", {})
                sans = extensions.get("subject_alternative_name", [])
                if sans:
                    result["ssl_san"] = ", ".join(sans)
                elif cert.get("subject", {}).get("CN"):
                    result["ssl_san"] = cert.get("subject", {}).get("CN")
                validity = cert.get("validity", {})
                if validity.get("not_before"):
                    result["ssl_valid_from"] = validity.get("not_before").split(" ")[0]

        res_resp = requests.get(VT_RELATIONS_URL.format(domain, "resolutions"), headers=headers, timeout=10)
        if res_resp.status_code == 200:
            for item in res_resp.json().get("data", []):
                match = re.search(r"(\d{1,3}(?:\.\d{1,3}){3})", item.get("id", ""))
                if match:
                    result["raw_hist_ips"].append(match.group(1))
            if result["raw_hist_ips"]:
                unique_hist_ips = list(dict.fromkeys(result["raw_hist_ips"]))
                result["raw_hist_ips"] = unique_hist_ips
                result["hist_ips_str"] = ", ".join(unique_hist_ips)

        ssl_resp = requests.get(VT_RELATIONS_URL.format(domain, "historical_ssl_certificates"), headers=headers,
                                timeout=10)
        if ssl_resp.status_code == 200:
            for item in ssl_resp.json().get("data", []):
                cert_id = item.get("id")
                if cert_id:
                    result["raw_hist_ssl"].append(cert_id)
            if result["raw_hist_ssl"]:
                unique_hist_ssl = list(dict.fromkeys(result["raw_hist_ssl"]))
                result["raw_hist_ssl"] = unique_hist_ssl
                result["hist_ssl_str"] = ", ".join(unique_hist_ssl)

        files_resp = requests.get(VT_RELATIONS_URL.format(domain, "communicating_files"), headers=headers, timeout=10)
        if files_resp.status_code == 200:
            rel_files = files_resp.json().get("data", [])
            hashes = [f.get("id") for f in rel_files if f.get("id")]
            if hashes:
                unique_hashes = list(dict.fromkeys(hashes))
                result["raw_hashes"] = unique_hashes
                result["vt_relations"] = ", ".join(unique_hashes)

    except Exception as e:
        print(f"[-] Ошибка VT API для {domain}: {e}")

    return result


def calculate_pair_weight(d1: dict, d2: dict) -> int:
    """Вычисление веса связности между доменами."""
    score = 0

    # 1. IP (текущие или исторические) -> 40 баллов
    ips1 = set(d1["raw_ips"] + d1["raw_hist_ips"])
    ips2 = set(d2["raw_ips"] + d2["raw_hist_ips"])
    if bool(ips1 & ips2):
        score += 40

    # 2. SSL-сертификаты -> 40 баллов
    certs1 = set(d1["raw_hist_ssl"]) | ({d1["ssl_sha256"]} if d1["ssl_sha256"] != "Н/Д" else set())
    certs2 = set(d2["raw_hist_ssl"]) | ({d2["ssl_sha256"]} if d2["ssl_sha256"] != "Н/Д" else set())
    if bool(certs1 & certs2):
        score += 40

    # 3. Файлы VT Relations -> 40 баллов
    hashes1 = set(d1["raw_hashes"])
    hashes2 = set(d2["raw_hashes"])
    if bool(hashes1 & hashes2):
        score += 40

    # 4. Общие NS или MX серверы -> 35 баллов (сильный маркер инфраструктуры)
    ns1 = set(d1["raw_ns"])
    ns2 = set(d2["raw_ns"])
    mx1 = set(d1["raw_mx"])
    mx2 = set(d2["raw_mx"])
    if bool(ns1 & ns2) or bool(mx1 & mx2):
        score += 35

    # 5. SOA RNAME (Email) -> 30 баллов
    soa1 = d1["soa_rname"]
    soa2 = d2["soa_rname"]
    if soa1 != "Н/Д" and soa1.lower() == soa2.lower():
        score += 30

    # 6. Подсети /24 (включая исторические IP) -> 20 баллов
    if not (ips1 & ips2):
        all_subs1 = {compute_subnet(ip) for ip in (d1["raw_ips"] + d1["raw_hist_ips"]) if compute_subnet(ip) != "Н/Д"}
        all_subs2 = {compute_subnet(ip) for ip in (d2["raw_ips"] + d2["raw_hist_ips"]) if compute_subnet(ip) != "Н/Д"}
        if bool(all_subs1 & all_subs2):
            score += 20

    # 7. Профиль (Регистратор + Страна с учетом эквивалентностей) -> 15 баллов
    def norm_c(c):
        c = str(c).strip().lower()
        if c in ['poland', 'pl']: return 'pl'
        if c in ['iceland', 'is']: return 'is'
        return c

    reg_match = (d1["registrar"] != "Н/Д" and d1["registrar"].lower() == d2["registrar"].lower())
    c_match = (d1["reg_country"] != "Н/Д" and norm_c(d1["reg_country"]) == norm_c(d2["reg_country"]))
    if reg_match and c_match:
        score += 15

    # 8. Пакетная регистрация (разница <= 2 дней) -> 15 баллов
    if d1["creation_dt"] and d2["creation_dt"]:
        delta_days = abs((d1["creation_dt"] - d2["creation_dt"]).days)
        if delta_days <= 2:
            score += 15

    return score


def build_weighted_clusters(domain_records: list, threshold: int = SCORE_THRESHOLD) -> tuple[dict, dict]:
    """Графовая кластеризация на базе весового порога."""
    adj = defaultdict(set)
    n = len(domain_records)

    for i in range(n):
        for j in range(i + 1, n):
            d1 = domain_records[i]
            d2 = domain_records[j]
            pair_score = calculate_pair_weight(d1, d2)

            if pair_score >= threshold:
                adj[d1["domain"]].add(d2["domain"])
                adj[d2["domain"]].add(d1["domain"])

    visited = set()
    clusters = []

    for dr in domain_records:
        dom = dr["domain"]
        if dom not in visited:
            component = []
            queue = [dom]
            visited.add(dom)

            while queue:
                curr = queue.pop(0)
                component.append(curr)
                for neighbor in adj[curr]:
                    if neighbor not in visited:
                        visited.add(neighbor)
                        queue.append(neighbor)
            clusters.append(component)

    multi_clusters = [c for c in clusters if len(c) > 1]
    single_clusters = [c for c in clusters if len(c) == 1]
    multi_clusters.sort(key=len, reverse=True)

    domain_to_cluster = {}
    cluster_color_map = {}

    for idx, comp in enumerate(multi_clusters, 1):
        c_name = f"Группа #{idx} ({len(comp)} домен.)"
        color_hex = CLUSTER_COLORS[(idx - 1) % len(CLUSTER_COLORS)]
        cluster_color_map[c_name] = color_hex
        for d in comp:
            domain_to_cluster[d] = (idx, c_name)

    for comp in single_clusters:
        domain_to_cluster[comp[0]] = (999, "Без связей")
        cluster_color_map["Без связей"] = None

    return domain_to_cluster, cluster_color_map


def main():
    parser = argparse.ArgumentParser(description="Анализ доменов с оптимизированной весовой кластеризацией")
    parser.add_argument("-f", "--file", required=True, help="Текстовый файл со списком доменов")
    parser.add_argument("-k", "--api-key", default="", help="VirusTotal API v3 ключ")
    parser.add_argument("-o", "--output", default="analysis_result.xlsx", help="Выходной файл XLSX")
    args = parser.parse_args()

    if not os.path.exists(args.file):
        print(f"[-] Файл {args.file} не найден!")
        sys.exit(1)

    with open(args.file, "r", encoding="utf-8") as f:
        domains = [clean_domain(line) for line in f if line.strip()]

    print(f"[+] Загружено {len(domains)} доменов для анализа...")

    domain_records = []
    for idx, d in enumerate(domains, 1):
        print(f"[{idx}/{len(domains)}] Анализируется: {d}")
        record = {"domain": d}
        vt_info = fetch_virustotal_data(d, args.api_key) if args.api_key else {
            "registrar": "Н/Д", "reg_country": "Н/Д", "creation_date": "Н/Д",
            "ns_records": "Н/Д", "mx_records": "Н/Д", "soa_rname": "Н/Д",
            "passive_ips": "Н/Д", "subnets": "Н/Д", "hist_ips_str": "Н/Д",
            "ssl_sha256": "Н/Д", "hist_ssl_str": "Н/Д", "ssl_san": "Н/Д",
            "ssl_valid_from": "Н/Д", "vt_relations": "Н/Д",
            "creation_dt": None,
            "raw_ns": [], "raw_mx": [], "raw_ips": [], "raw_subnets": [],
            "raw_hist_ips": [], "raw_hist_ssl": [], "raw_hashes": []
        }
        record.update(vt_info)
        domain_records.append(record)

    # 1. Весовая кластеризация
    domain_to_cluster, cluster_color_map = build_weighted_clusters(domain_records, threshold=SCORE_THRESHOLD)

    for dr in domain_records:
        c_order, c_label = domain_to_cluster[dr["domain"]]
        dr["cluster_order"] = c_order
        dr["cluster_label"] = c_label

    domain_records.sort(key=lambda x: (x["cluster_order"], x["domain"]))

    # 2. Подсчет общих признаков для точечной подсветки ячеек
    domain_count_by = {
        "ns": Counter(), "mx": Counter(),
        "ip": Counter(), "subnet": Counter(),
        "hash": Counter(), "registrar": Counter(), "country": Counter(),
        "soa": Counter(), "ssl_sha": Counter(), "creation_date": Counter()
    }

    for dr in domain_records:
        for ns in set(dr["raw_ns"]): domain_count_by["ns"][ns] += 1
        for mx in set(dr["raw_mx"]): domain_count_by["mx"][mx] += 1
        for h in set(dr["raw_hashes"]): domain_count_by["hash"][h] += 1
        for sub in set(dr["raw_subnets"]): domain_count_by["subnet"][sub] += 1

        for ip in set(dr["raw_ips"] + dr["raw_hist_ips"]):
            domain_count_by["ip"][ip] += 1

        all_certs = set(dr["raw_hist_ssl"])
        if dr["ssl_sha256"] != "Н/Д":
            all_certs.add(dr["ssl_sha256"])
        for cert in all_certs:
            domain_count_by["ssl_sha"][cert] += 1

        if dr["registrar"] != "Н/Д": domain_count_by["registrar"][dr["registrar"].lower()] += 1
        if dr["reg_country"] != "Н/Д": domain_count_by["country"][dr["reg_country"].lower()] += 1
        if dr["soa_rname"] != "Н/Д": domain_count_by["soa"][dr["soa_rname"].lower()] += 1
        if dr["creation_date"] != "Н/Д": domain_count_by["creation_date"][dr["creation_date"]] += 1

    shared = {k: {item for item, count in v.items() if count > 1} for k, v in domain_count_by.items()}

    # 3. Сохранение в Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IoC Анализ доменов"

    headers = [
        "Кластер (Группа)", "Домен", "Регистратор", "Страна регистранта",
        "Дата создания", "NS записи", "MX записи", "SOA RNAME",
        "Пассивные IP (A)", "Исторические IP", "Подсеть /24",
        "Текущий SSL SHA-256", "Исторические SSL", "SSL SAN",
        "SSL действует с", "Связанные файлы (VT Relations)"
    ]

    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                         top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

    ws.append(headers)
    for col_num, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for row_idx, dr in enumerate(domain_records, 2):
        row_data = [
            dr["cluster_label"], defang_domain(dr["domain"]), dr["registrar"], dr["reg_country"],
            dr["creation_date"], dr["ns_records"], dr["mx_records"], dr["soa_rname"],
            dr["passive_ips"], dr["hist_ips_str"], dr["subnets"],
            dr["ssl_sha256"], dr["hist_ssl_str"], dr["ssl_san"],
            dr["ssl_valid_from"], dr["vt_relations"]
        ]
        ws.append(row_data)

        c_label = dr["cluster_label"]
        cluster_color_hex = cluster_color_map.get(c_label)
        row_highlight_fill = PatternFill(start_color=cluster_color_hex, end_color=cluster_color_hex,
                                         fill_type="solid") if cluster_color_hex else None

        highlights = [
            bool(cluster_color_hex),
            False,
            dr["registrar"].lower() in shared["registrar"],
            dr["reg_country"].lower() in shared["country"],
            dr["creation_date"] in shared["creation_date"],
            any(ns in shared["ns"] for ns in dr["raw_ns"]),
            any(mx in shared["mx"] for mx in dr["raw_mx"]),
            dr["soa_rname"].lower() in shared["soa"],
            any(ip in shared["ip"] for ip in dr["raw_ips"]),
            any(ip in shared["ip"] for ip in dr["raw_hist_ips"]),
            any(sub in shared["subnet"] for sub in dr["raw_subnets"]),
            dr["ssl_sha256"] in shared["ssl_sha"],
            any(cert in shared["ssl_sha"] for cert in dr["raw_hist_ssl"]),
            False,
            False,
            any(h in shared["hash"] for h in dr["raw_hashes"])
        ]

        for col_idx, is_highlighted in enumerate(highlights, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="left" if col_idx in [1, 2, 3, 6, 7, 8, 9, 10, 11, 12, 13, 14, 16] else "center",
                vertical="center")

            if is_highlighted and row_highlight_fill:
                cell.fill = row_highlight_fill

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=15)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 4, 15), 45)

    wb.save(args.output)
    print(f"\n[+] Анализ завершен! Результат сохранен в: {args.output}")


if __name__ == "__main__":
    main()